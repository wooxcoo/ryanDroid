import openpyxl as xl
from time import sleep
import pandas as pd
# import pyodbc
print("\r This code is created for PR4053 Surazh PM by Ryan You from Spin AU on 15-DEC-2019;")
print("\r Steps:  1 Export blocks in POV of PCS7 ;")
print("\r         2 Copy the CPU21_B.csv to the same folder as the app;")
print("\r         3 Copy the diagnosis.mdb to the same folder as the app;")
print("\r         4 Run the app to add new data into the database for Wincc;")
print("\r It is only used to export the Loopspy database from CPU21 CFC blocks;")
print("\r Please wait and keep the window open until 100% is done.")
print("\r If the process stuck at one point, please try to run the app again.")
print("\r Starting...")
read_file = pd.read_csv(r'CPU21_B.csv', encoding='cp1252')
read_file.to_excel(r'Loopspy.xlsx', index = None, header=True)
wb = xl.load_workbook('Loopspy.xlsx')
sheet = wb['Sheet1']
sheet.title = 'CPU21_B'
wb.save('Loopspy.xlsx')
new_sheet = wb.create_sheet('LoopSpy')
new_sheet.cell(1, 1).value = 'SignalNumber'
new_sheet.cell(1, 2).value = 'VOITHSignalNumber'
new_sheet.cell(1, 3).value = 'Area'
new_sheet.cell(1, 4).value = 'Language1'
new_sheet.cell(1, 5).value = 'Language2'
new_sheet.cell(1, 6).value = 'Language3'
new_sheet.cell(1, 7).value = 'Language4'
new_sheet.cell(1, 8).value = 'Language5'

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)  # block
    text = str(cell.value)
    first = str(text[:1])
    second = str(text[:2])
    third = str(text[:3])
    row1 = new_sheet.max_row
    p = 100*row//sheet.max_row
    print("\r[%-100s] %d%%" % ('>' * p, 1 * p), end='')
    sleep(0)
    if first == 'M' and second != 'MO' and second != 'MU' and second != 'MS' and second != 'MA' \
            and second != 'MI' and second != 'MM' and second != 'Mo' and second != 'MW':
        cell2 = sheet.cell(row, 2)
        cell3 = \
            sheet.cell(row, 3)
        row1 = new_sheet.max_row + 1
        loop_cell = new_sheet.cell(row1, 1)
        lang1_cell = new_sheet.cell(row1, 4)
        lang2_cell = new_sheet.cell(row1, 5)
        loop_tag = str(cell2.value) + '.' + str(cell3.value)
        loop_cell.value = loop_tag
        lang1_cell.value = sheet.cell(row, 4).value
        lang2_cell.value = sheet.cell(row, 4).value
    elif first == '0':
        cell2 = sheet.cell(row, 2)
        cell3 = sheet.cell(row, 3)
        row1 = new_sheet.max_row + 1
        loop_cell = new_sheet.cell(row1, 1)
        lang1_cell = new_sheet.cell(row1, 4)
        lang2_cell = new_sheet.cell(row1, 5)
        loop_tag = str(cell2.value) + '.' + str(cell3.value)
        loop_cell.value = loop_tag
        lang1_cell.value = sheet.cell(row, 4).value
        lang2_cell.value = sheet.cell(row, 4).value
    elif first == 'X':
        cell2 = sheet.cell(row, 2)
        cell3 = sheet.cell(row, 3)
        row1 = new_sheet.max_row + 1
        loop_cell = new_sheet.cell(row1, 1)
        lang1_cell = new_sheet.cell(row1, 4)
        lang2_cell = new_sheet.cell(row1, 5)
        loop_tag = str(cell2.value) + '.' + str(cell3.value)
        loop_cell.value = loop_tag
        lang1_cell.value = sheet.cell(row, 4).value
        lang2_cell.value = sheet.cell(row, 4).value
        loop_cell_1 = new_sheet.cell(row1 + 1, 1)
        lang1_cell_1= new_sheet.cell(row1 + 1, 4)
        lang2_cell_1 = new_sheet.cell(row1 +1, 5)
        loop_tag_1 = str(cell2.value) + '/' + str(cell3.value)
        loop_cell_1.value = loop_tag_1
        lang1_cell_1.value = sheet.cell(row, 4).value
        lang2_cell_1.value = sheet.cell(row, 4).value
    else:
        pass
print("\r")
wb.save('Loopspy.xlsx')
print("\r Congralations!The data has been created in a new sheet of Loopspy.xlsx.")
sleep(5)
# conn_str = (
#     r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
#     r'DBQ=C:\Users\Ryan\Downloads\surazh\surazh/Diagnose.mdb;'
#     )
# cnxn = pyodbc.connect(conn_str)
# crsr = cnxn.cursor()
# crsr.execute('''
#                    INSERT INTO names_table (SignalNumber, Lanuage0, Lanuage1)
#                    VALUES('Mia', 'Mogran',66)
#              ''')
#
# conn_str.commit()
import openpyxl as xl
from time import sleep, strftime, localtime
from app.model.loopspy import LoopspyTable
import datetime

import pandas as pd
import json
# import pyodbc


def printTime(step_name, last_time=None):

    now_time = datetime.datetime.now()
    if last_time:
        print ( "last step cost: {}\n\n".format( now_time - last_time))

    print("begin {} at {} \n".format(step_name, strftime("%Y-%m-%d %H:%M:%S", localtime())))
    return now_time


def print_help():
    print("\r This code is created for PR4053 Surazh PM by Ryan You from Spin AU on 15-DEC-2019;")
    print("\r Steps:  1 Export blocks in POV of PCS7 ;")
    print("\r         2 Copy the CPU21_B.csv to the same folder as the app;")
    print("\r         3 Copy the diagnosis.mdb to the same folder as the app;")
    print("\r         4 Run the app to add new data into the database for Wincc;")
    print("\r It is only used to export the Loopspy database from CPU21 CFC blocks;")
    print("\r Please wait and keep the window open until 100% is done.")
    print("\r If the process stuck at one point, please try to run the app again.")
    print("\r Starting...")


# Hierarchy	Chart	Block	Block comment	Create block icon	Block icon	OCM possible	Readback allowed	Block group	Block type	Unnamed: 10
 # 先人肉写死
DATA_MAPPING = {
    "hierarchy": 0,  # Hierarchy
    "chart": 1,  # Chart
    "block": 2,  # Block
    "block_comment": 3, # Block comment
    "create_block_icon": 4,  # Create block icon
    "block_icon": 5,   # Block icon
    "ocm_possible": 6,  # OCM possible
    "readback_allowed": 7,  # Readback allowed
    "block_group": 8,   # Block group
    "block_type": 9   # Block type
}

# xsl mapping
def get_xsl_struecture():
    # todo: 第一行是 row 表头， 从表头中获取列名， 存储 mapping from column_name to index
    return DATA_MAPPING

def get_xsl_to_dict(row_item, data_mapping=DATA_MAPPING):
    """
    get dict data from xsl row_item
    :param row_item:
    :return:
    """
    ret_data = {}
    for attr_keuy, index in data_mapping.items():
        ret_data[attr_keuy] = row_item[index].value
    return ret_data

# 把excel 表遍历， 内容读取， 条件判断 和 数据入sql 库分开写， 可以把复杂逻辑拆成简单的步骤
# 条件判断和入库逻辑应该是强耦合的，可以写到一起
# 如果有多种入库方法，分别独立成小函数；以后每次就改对应函数就可以
# 要是有时间，把 excel 表 和 sql 表的对应关系抽象成配置，，工具就可以通用了， 以后改下 对应关系，就可以处理多种情况，

def parse_logic(row_dict, data_mapping=DATA_MAPPING, proc_func=None):
    # 把excel 的处理逻辑集中到这里
    # 明确需求逻辑 block
    # case sensitive？
    # condition 1, upper case?, statrt with M, but not with ['MO', 'MU', 'MS', 'MA', 'MI', 'MM', 'MW']
    # condition 2, start with 0
    # condition 3, start with x
    # ignore others
    #


    block = row_dict.get("block")
    SignalNumber = "{}.{}".format(row_dict.get("chart"), row_dict.get("block"))
    Language1 = row_dict.get("block_comment")

    # 入库数据
    db_item = row_dict
    db_item["SignalNumber"] = SignalNumber
    db_item["Language1"] = Language1

    # todo：全部入 sql 库， igonre 的item valid=false
    # 通过 block 字段判断
    # if first == 'M' and second != 'MO' and second != 'MU' and second != 'MS' and second != 'MA' \
    #        and second != 'MI' and second != 'MM' and second != 'Mo' and second != 'MW':
    # elif first == '0':
    # elif first == 'X':
    # ignore others
    valid = False
    if block.startswith('0') or block.startswith('X'):
        valid = True
    elif block.startswith('M'):
        if len(block) == 1:
            # 只有一个字母
            valid = True
        else:
            first_two = block[:2]
            if first_two not in ["MO", "MU", "MS", "MA", "MI", "MM", "Mo", "MW"]:
                valid = True

    db_item["valid"] = valid
    return db_item, valid

def store_data(db_data):
    # todo: 通用处理， 更具具体逻辑，对应写入表
    pass


def main_logic(inpu_file=r'CPU21_B.csv', encoding=r'cp1252',  out_put_name=r'Loopspy_v1.xlsx'):
    # in the same folder

    c_time = printTime("begin_read")
    read_file = pd.read_csv(inpu_file, encoding=encoding)
    c_time = printTime("pd.read_csv", c_time)
    read_file.to_excel(out_put_name, index = None, header=True)
    wb = xl.load_workbook(out_put_name)
    c_time = printTime("load pd to excel", c_time)
    # no need to convert csv to xlsx;
    sheet = wb['Sheet1']
    sheet.title = 'CPU21_B'

    # 加一个新 sheet， 表头
    new_sheet = wb.create_sheet('LoopSpy')
    new_sheet.cell(1, 1).value = 'SignalNumber'
    new_sheet.cell(1, 2).value = 'VOITHSignalNumber'
    new_sheet.cell(1, 3).value = 'Area'
    new_sheet.cell(1, 4).value = 'Language1'
    new_sheet.cell(1, 5).value = 'Language2'
    new_sheet.cell(1, 6).value = 'Language3'
    new_sheet.cell(1, 7).value = 'Language4'
    new_sheet.cell(1, 8).value = 'Language5'

    # model definition
    # SignalNumber,  VOITHSignalNumber,  Area,   Language[1-5] ?

    # 读到字典里， 一次写db?
    iter_rows = sheet.iter_rows()
    # 把头读出来不处理
    header_row = next(iter_rows)
    ret_data = get_xsl_to_dict(header_row)
    print (json.dumps(ret_data))

    cnt = 1
    for row_item in iter_rows:
        rew_dict = get_xsl_to_dict(row_item)
        # print (json.dumps(rew_dict))
        p = 100*cnt//sheet.max_row
        cnt += 1
        print("\r[%-100s] %d%%" % ('>' * p, 1 * p), end='')
        sleep(0)
        db_item, valid = parse_logic(rew_dict)
        # db write
        LoopspyTable.insert_one(**db_item)
        if valid:
            # new sheet
            row1 = new_sheet.max_row + 1
            loop_cell = new_sheet.cell(row1, 1)
            lang1_cell = new_sheet.cell(row1, 4)
            lang2_cell = new_sheet.cell(row1, 5)
            loop_cell.value = db_item.get("SignalNumber")
            lang1_cell.value = db_item.get("block_comment")
            lang2_cell.value = db_item.get("block_comment")

    print("\r")
    c_time = printTime("parse_ready", c_time)
    wb.save(out_put_name)
    printTime("save to new sheet", c_time)
    print("\r Congralations!The data has been created in a new sheet of Loopspy.xlsx.")


# unit test
if __name__ == '__main__':
    print_help()
    main_logic()
